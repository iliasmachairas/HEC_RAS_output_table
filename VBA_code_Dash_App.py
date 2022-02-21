# -*- coding: utf-8 -*-
"""
Created on Sun Feb 20 09:59:40 2022

@author: ilias
"""

import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pandas as pd
import datetime
from io import BytesIO

import os
import shutil
import numpy as np

import win32com.client
from pywintypes import com_error

st.set_page_config(layout = "wide")

st.markdown("""# HEC-RAS output from .txt to pdf""")

col2, space2, col3 = st.columns((10,1,10))

with col2:
    # Read data
    upload_data = st.file_uploader('Select the file', type=['txt'])
    if upload_data is not None:
        data = pd.read_fwf(upload_data, skiprows=10)
        col_names_English = ['Reach', 'River Station', 'Profile', 'Number', 'Q Total', 'Min Ch El', 'W.S. Elev', 'Crit W.S.',
                    'E.G. Elev', 'E.G. Slope', 'Vel Chnl', 'Flow Area', 'Top Width', 'Froude # Chl']
        data.columns = col_names_English
        #st.write(data.head())
    
        # data wrangling
        data['Number'] = data['Number'].astype('str')
        data['Profile'] = data['Profile'] + data['Number']
        data.drop('Number', axis=1, inplace=True)
        data.head()
        
        # Plots
        data_ploting = data.copy()
        data_ploting.set_index('River Station', inplace=True)
        fig = make_subplots(rows=2, cols=1)
        trace_1 = go.Scatter(x=data_ploting.index, y=data_ploting['W.S. Elev'], name='water level (m)')
        trace_2 = go.Scatter(x=data_ploting.index, y=data_ploting['Min Ch El'], name='ground surface (m)')
        trace_3 = go.Scatter(x=data_ploting.index, y=data_ploting['Vel Chnl'], name='velocity (m/s)')
        data_subplot_1 = [trace_1, trace_2]
        fig.append_trace(trace_1, row=1, col=1)
        fig.append_trace(trace_2, row=1, col=1)
        fig.append_trace(trace_3, row=2, col=1)
        fig.update_layout(height=600, width=600, title_text="Hydraulic charactersistics")
        st.plotly_chart(fig, use_container_width=True)
        
        # Creating directory

    
        cwd = os.getcwd()
        direcory_name = 'HEC_RAS_excel_pdf'
        path = os.path.join(cwd, direcory_name)
        
        # if os.path.exists(path) and os.path.isdir(path):
        #     shutil.rmtree(path)
        # os.mkdir(path)
        
        excel_name = os.path.join(path, 'Hydraulic.xlsx')
        print(excel_name)
        data.to_excel(excel_name)
        
        # Excel manipulation

        import openpyxl
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Border, Side
        
        wb = load_workbook('HEC_RAS_excel_pdf/Hydraulic.xlsx')
        ws = wb['Sheet1']
        
        # Adjusting column width size
        column_widths = []
        for row in wb['Sheet1'].iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))
        
                    
        col_width_array = np.asarray(column_widths, dtype=np.float32)        
        #print(column_widths)
        columns_length = len(column_widths)
        col_width_array =  col_width_array + 5
        #print(col_width_array)
                    
        for i, column_width in enumerate(col_width_array):
            wb['Sheet1'].column_dimensions[get_column_letter(i + 1)].width = column_width
            
        ws.insert_rows(1)
        # Freezing top rows
        ws.freeze_panes = "O4"
        
        # Inserting a row
        ws.insert_rows(3)
        units_row = ['(m3/s)', '(m)', '(m)', '(m)',  '(m)', '(m/m)',  '(m/s)',  '(m2)', '(m)', '-']
        
        for v,c in zip(units_row, range(5,15)):
            ws.cell(column=c, row=3).value = v
        
        # Fixing alignment
        from openpyxl.styles.alignment import Alignment
        last_row = len(list(ws.rows))
        last_column = ws.max_column
        
        for j in range(1,last_column+1):
            for i in range(2,last_row+1):
                ws.cell(column=j, row=i).alignment = Alignment(horizontal='center')
        
        # applying border style
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for j in range(1,last_column+1):
            for i in range(2,last_row+1):
                ws.cell(column=j, row=i).border = thin_border
        
        # fixing superscripts
        ws['E3'] = '(m'+ u'\u00B3' +'/s)'
        ws['L3'] ='(m'+ u'\u00B2'+')'
        
        # Printing settings
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(ws, paper_size = 9, orientation='landscape')
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.print_title_rows = '1:3'
        wb.save('HEC_RAS_excel_pdf/Hydraulic.xlsx')
        
        # Printing to pdf
        excel = win32com.client.Dispatch("Excel.Application")
          
        # Read Excel File
        excel.Visible = False
        excel.DisplayAlerts=False
        sheets = excel.Workbooks.Open(excel_name)
        # sheets('Sheet1').PageSetup.PrintTitleRows = "$1:$2"
        
        path_pdf = os.path.join(path, 'Hydraulic_Calc.pdf')
        
        sheets.ExportAsFixedFormat(Type=0,  Filename=path_pdf,
                                  IgnorePrintAreas=False)
        
        with col3:
            selected_cat = st.selectbox(label='Select parameter', options=['Only Excel file','Only Pdf file','Both'])
            
            with open('HEC_RAS_excel_pdf/Hydraulic.xlsx', "rb") as file:
                btn=st.download_button(
                label="click me to download EXCEL",
                data=file,
                file_name="dowloaded_EXCEL.xlsx",
                mime="application/vnd.ms-excel"
                )
                
            # add a zip file - create it via Python first
            
            # with open(path_pdf, "rb") as file:
            #     btn=st.download_button(
            #     label="click me to download pdf",
            #     data=file,
            #     file_name="dowloaded.pdf",
            #     mime="application/octet-stream"
            #     )
            
            # if selected_cat == 'Only Excel file':
            #     st.download_button(label = "Download data as Excel", data = 'HEC_RAS_excel_pdf/Hydraulic.xlsx', file_name = 'Hydraulic.xlsx')
            # elif selected_cat == 'Only Pdf file':
            #     st.download_button(label = "Download data as Excel", data = 'HEC_RAS_excel_pdf/Hydraulic_Calc.pdf', file_name = 'Hydraulic_Calc.pdf')
            # #else:
            #     #st.download_button(

    
    